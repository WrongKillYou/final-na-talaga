# monitoring/views.py
# CLEANED AND CORRECTED FOR KINDERGARTEN COMPETENCY-BASED SYSTEM

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Count, Q, Avg
from django.core.paginator import Paginator
from django.utils import timezone
from datetime import date, timedelta, datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from users.models import Child
from users.decorators import teacher_required, parent_required
from .models import (
    Class, Enrollment, Attendance, AttendanceSummary,
    Domain, Competency, QuarterlyCompetencyRecord, QuarterlySummary
)


# ========================================
# HELPER FUNCTIONS
# ========================================

def get_current_quarter():
    """
    Determine current quarter based on Philippine school calendar
    Adjust dates as needed for your school
    """
    month = datetime.now().month
    
    if month in [6, 7, 8]:
        return 1  # 1st Quarter (June-August)
    elif month in [9, 10, 11]:
        return 2  # 2nd Quarter (September-November)
    elif month in [12, 1, 2]:
        return 3  # 3rd Quarter (December-February)
    else:  # [3, 4, 5]
        return 4  # 4th Quarter (March-May)


def get_attendance_stats(child, start_date=None, end_date=None, class_obj=None):
    """
    Calculate attendance statistics for a child
    """
    if not end_date:
        end_date = date.today()
    
    if not start_date:
        start_date = end_date.replace(day=1)
    
    # Query attendance records
    attendance_records = Attendance.objects.filter(
        child=child,
        date__gte=start_date,
        date__lte=end_date
    )
    
    if class_obj:
        attendance_records = attendance_records.filter(class_obj=class_obj)
    
    # Count different statuses
    total_days = attendance_records.count()
    present_days = attendance_records.filter(status='present').count()
    absent_days = attendance_records.filter(status='absent').count()
    late_days = attendance_records.filter(status='late').count()
    excused_days = attendance_records.filter(status='excused').count()
    
    # Calculate attendance rate
    attendance_rate = 0
    if total_days > 0:
        attendance_rate = round((present_days / total_days) * 100, 1)
    
    return {
        'total_days': total_days,
        'present_days': present_days,
        'absent_days': absent_days,
        'late_days': late_days,
        'excused_days': excused_days,
        'attendance_rate': attendance_rate,
    }


# ========================================
# CLASS MANAGEMENT
# ========================================

@login_required
@teacher_required
def class_list(request):
    """Display all classes assigned to the logged-in teacher"""
    teacher = request.user.teacher_profile
    
    classes = Class.objects.filter(
        teacher=teacher,
        is_active=True
    ).annotate(
        student_count=Count('enrollments')
    ).order_by('-school_year', 'class_name')
    
    context = {
        'teacher': teacher,
        'classes': classes,
    }
    return render(request, 'monitoring/class_list.html', context)


@login_required
@teacher_required
def class_detail(request, class_id):
    """Show details of a class and enrolled students"""
    teacher = request.user.teacher_profile
    class_obj = get_object_or_404(
        Class, 
        id=class_id, 
        teacher=teacher
    )
    
    # Get all enrolled students
    enrollments = Enrollment.objects.filter(
        class_obj=class_obj,
        is_active=True
    ).select_related('student').order_by(
        'student__last_name', 
        'student__first_name'
    )
    
    # Get current quarter
    current_quarter = get_current_quarter()
    
    # Prepare student data with statistics
    students_data = []
    for enrollment in enrollments:
        student = enrollment.student
        
        # Get competency records for current quarter
        records = QuarterlyCompetencyRecord.objects.filter(
            child=student,
            quarter=current_quarter
        )
        
        # Count competency levels
        total_assessed = records.exclude(level__isnull=True).count()
        consistent_count = records.filter(level='C').count()
        
        # Get attendance stats (current month)
        attendance_stats = get_attendance_stats(student, class_obj=class_obj)
        
        students_data.append({
            'student': student,
            'enrollment': enrollment,
            'total_assessed': total_assessed,
            'consistent_count': consistent_count,
            'attendance_rate': attendance_stats['attendance_rate'],
        })
    
    # Class statistics
    total_students = len(students_data)
    avg_attendance = sum(s['attendance_rate'] for s in students_data) / total_students if total_students else 0
    
    context = {
        'teacher': teacher,
        'class_obj': class_obj,
        'students_data': students_data,
        'total_students': total_students,
        'avg_attendance': round(avg_attendance, 1),
        'current_quarter': current_quarter,
    }
    
    return render(request, 'monitoring/class_detail.html', context)


# ========================================
# COMPETENCY RECORD MANAGEMENT
# ========================================

@login_required
@teacher_required
def competency_input(request, class_id):
    """Input competency records for students in a class"""
    teacher = request.user.teacher_profile
    class_obj = get_object_or_404(Class, id=class_id, teacher=teacher)
    
    # Get selected quarter from GET params or default to current
    selected_quarter = int(request.GET.get('quarter', get_current_quarter()))
    
    # Get all domains and competencies
    domains = Domain.objects.prefetch_related('competencies').all()
    
    # Get enrolled students
    enrollments = Enrollment.objects.filter(
        class_obj=class_obj,
        is_active=True
    ).select_related('student').order_by(
        'student__last_name', 
        'student__first_name'
    )
    
    context = {
        'teacher': teacher,
        'class_obj': class_obj,
        'selected_quarter': selected_quarter,
        'quarters': [1, 2, 3, 4],
        'domains': domains,
        'enrollments': enrollments,
    }
    
    return render(request, 'monitoring/competency_input.html', context)


@login_required
@teacher_required
def download_competency_template(request, class_id):
    """
    Download Excel template for competency record entry
    Pre-filled with student names and competency list
    """
    teacher = request.user.teacher_profile
    class_obj = get_object_or_404(Class, id=class_id, teacher=teacher)
    
    quarter = int(request.GET.get('quarter', get_current_quarter()))
    
    # Get enrolled students
    enrollments = Enrollment.objects.filter(
        class_obj=class_obj,
        is_active=True
    ).select_related('student').order_by(
        'student__last_name',
        'student__first_name'
    )
    
    # Get all competencies organized by domain
    domains = Domain.objects.prefetch_related('competencies').all()
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Q{quarter} Records"
    
    # Styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title row
    ws.merge_cells('A1:F1')
    ws['A1'] = f"COMPETENCY RECORD TEMPLATE - {class_obj.class_name} - Quarter {quarter}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Instructions row
    ws.merge_cells('A2:F2')
    ws['A2'] = "Enter B (Beginning), D (Developing), or C (Consistent) for each competency. Leave blank if not assessed."
    ws['A2'].font = Font(italic=True, size=10)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Header row
    ws.append([])  # Empty row
    headers = ['Student LRN', 'Student Name', 'Domain', 'Competency Code', 'Competency Description', 'Level (B/D/C)']
    ws.append(headers)
    
    header_row = ws[4]
    for cell in header_row:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 60
    ws.column_dimensions['F'].width = 12
    
    # Fill data rows
    row_num = 5
    for enrollment in enrollments:
        child = enrollment.student
        
        for domain in domains:
            for competency in domain.competencies.all():
                # Get existing record if any
                try:
                    record = QuarterlyCompetencyRecord.objects.get(
                        child=child,
                        competency=competency,
                        quarter=quarter
                    )
                    level = record.level or ''
                except QuarterlyCompetencyRecord.DoesNotExist:
                    level = ''
                
                ws.append([
                    child.get_full_name,
                    child.get_full_name(),
                    domain.name,
                    competency.code,
                    competency.description,
                    level
                ])
                
                # Apply border to all cells
                for cell in ws[row_num]:
                    cell.border = border
                    if cell.column == 6:  # Level column
                        cell.alignment = Alignment(horizontal='center')
                
                row_num += 1
    
    # Freeze panes
    ws.freeze_panes = 'A5'
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="Competency_Template_{class_obj.class_name}_Q{quarter}.xlsx"'
    )
    wb.save(response)
    return response


@login_required
@teacher_required
def bulk_upload_competencies(request, class_id):
    """
    Upload filled Excel template to bulk update competency records
    """
    teacher = request.user.teacher_profile
    class_obj = get_object_or_404(Class, id=class_id, teacher=teacher)
    
    if request.method != 'POST':
        messages.error(request, 'Invalid request method.')
        return redirect('monitoring:competency_input', class_id=class_id)
    
    excel_file = request.FILES.get('competency_file')
    if not excel_file:
        messages.error(request, 'No file uploaded.')
        return redirect('monitoring:competency_input', class_id=class_id)
    
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
        
        success_count = 0
        error_count = 0
        errors = []
        
        # Process rows (skip header rows - start from row 5)
        for row_num, row in enumerate(ws.iter_rows(min_row=5, values_only=True), start=5):
            try:
                lrn, student_name, domain_name, comp_code, comp_desc, level = row[:6]
                
                # Skip if LRN is empty
                if not lrn:
                    continue
                
                # Validate level
                if level and level.upper() not in ['B', 'D', 'C']:
                    errors.append(f"Row {row_num}: Invalid level '{level}' (must be B, D, or C)")
                    error_count += 1
                    continue
                
                # Get student
                try:
                    student = Child.objects.get(lrn=lrn)
                except Child.DoesNotExist:
                    errors.append(f"Row {row_num}: Student with LRN {lrn} not found")
                    error_count += 1
                    continue
                
                # Verify enrollment
                if not Enrollment.objects.filter(
                    student=student, 
                    class_obj=class_obj,
                    is_active=True
                ).exists():
                    errors.append(f"Row {row_num}: {student.get_full_name()} not enrolled in this class")
                    error_count += 1
                    continue
                
                # Get competency
                try:
                    competency = Competency.objects.get(code=comp_code)
                except Competency.DoesNotExist:
                    errors.append(f"Row {row_num}: Competency {comp_code} not found")
                    error_count += 1
                    continue
                
                # Extract quarter from filename or use current
                quarter = get_current_quarter()
                
                # Create or update record
                if level:
                    record, created = QuarterlyCompetencyRecord.objects.update_or_create(
                        child=student,
                        competency=competency,
                        quarter=quarter,
                        defaults={
                            'level': level.upper(),
                            'recorded_by': teacher
                        }
                    )
                    success_count += 1
                
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
                error_count += 1
                continue
        
        # Show results
        if success_count > 0:
            messages.success(request, f'✅ Successfully uploaded {success_count} competency record(s).')
        
        if error_count > 0:
            error_message = f'⚠️ {error_count} error(s) encountered. '
            if errors:
                error_message += 'First 5 errors: ' + '; '.join(errors[:5])
            messages.warning(request, error_message)
        
        if success_count == 0 and error_count == 0:
            messages.info(request, 'No records were processed.')
        
    except Exception as e:
        messages.error(request, f'❌ Error processing file: {str(e)}')
    
    return redirect('monitoring:competency_input', class_id=class_id)


@login_required
@teacher_required
def student_competency_detail(request, student_id):
    """View detailed competency records for a specific student"""
    teacher = request.user.teacher_profile
    student = get_object_or_404(Child, id=student_id)
    
    # Verify teacher has access to this student
    if not Enrollment.objects.filter(
        student=student,
        class_obj__teacher=teacher,
        is_active=True
    ).exists():
        messages.error(request, 'You do not have access to this student.')
        return redirect('monitoring:class_list')
    
    # Get all competency records grouped by quarter and domain
    quarters_data = []
    for quarter in [1, 2, 3, 4]:
        domains_data = []
        
        for domain in Domain.objects.prefetch_related('competencies').all():
            competencies_data = []
            
            for competency in domain.competencies.all():
                try:
                    record = QuarterlyCompetencyRecord.objects.get(
                        child=student,
                        competency=competency,
                        quarter=quarter
                    )
                    level = record.level
                    notes = record.notes
                except QuarterlyCompetencyRecord.DoesNotExist:
                    level = None
                    notes = ''
                
                competencies_data.append({
                    'competency': competency,
                    'level': level,
                    'notes': notes,
                })
            
            domains_data.append({
                'domain': domain,
                'competencies': competencies_data,
            })
        
        quarters_data.append({
            'quarter': quarter,
            'domains': domains_data,
        })
    
    context = {
        'teacher': teacher,
        'student': student,
        'quarters_data': quarters_data,
    }
    
    return render(request, 'monitoring/student_competency_detail.html', context)


# ========================================
# ATTENDANCE MANAGEMENT
# ========================================
from django.core.paginator import Paginator

@login_required
@teacher_required
def attendance_list(request):
    teacher = request.user.teacher_profile
    classes = Class.objects.filter(teacher=teacher, is_active=True)
    
    class_id = request.GET.get('class_id')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    status = request.GET.get('status')
    
    # Base query
    attendance_records = Attendance.objects.filter(
        class_obj__teacher=teacher
    ).select_related('child', 'class_obj', 'recorded_by').order_by('-date', 'child__last_name')

    # Apply filters
    if class_id:
        attendance_records = attendance_records.filter(class_obj_id=class_id)
    if start_date:
        attendance_records = attendance_records.filter(date__gte=start_date)
    if end_date:
        attendance_records = attendance_records.filter(date__lte=end_date)
    if status:
        attendance_records = attendance_records.filter(status=status)

    # Pagination
    paginator = Paginator(attendance_records, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'teacher': teacher,
        'classes': classes,
        'page_obj': page_obj,
        'status_choices': Attendance.STATUS_CHOICES,
    }

    return render(request, 'monitoring/attendance_list.html', context)


# monitoring/views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from datetime import date, datetime
from users.models import Teacher, Child
from .models import Attendance
from users.decorators import teacher_required


@login_required
@teacher_required
def record_attendance(request):
    """Page for teachers to record attendance for their class"""
    teacher = request.user.teacher_profile
    classes = Class.objects.filter(teacher=teacher, is_active=True)

    # Get selected date and class
    selected_date_str = request.GET.get('date', date.today().isoformat())
    selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
    selected_class_id = request.GET.get('class_id')
    selected_class = None
    children = []

    if selected_class_id:
        selected_class = get_object_or_404(Class, id=selected_class_id, teacher=teacher)
        enrollments = Enrollment.objects.filter(
            class_obj=selected_class,
            is_active=True
        ).select_related('student').order_by('student__last_name', 'student__first_name')

        # Get existing attendance records for that date
        existing_records = {
            record.child_id: record
            for record in Attendance.objects.filter(
                date=selected_date,
                class_obj=selected_class
            )
        }

        # Build children list for template
        for enrollment in enrollments:
            child = enrollment.student
            record = existing_records.get(child.id)
            children.append({
                'child': child,
                'status': record.status if record else None,
                'time_in': record.time_in if record else None,
                'time_out': record.time_out if record else None,
                'remarks': record.remarks if record else '',
            })

    # Handle form submission
    if request.method == 'POST':
        form_date = datetime.strptime(request.POST.get('date'), '%Y-%m-%d').date()
        form_class_id = request.POST.get('class_id')
        form_class = get_object_or_404(Class, id=form_class_id, teacher=teacher)

        saved_count = 0
        enrollments = Enrollment.objects.filter(
            class_obj=form_class,
            is_active=True
        ).select_related('student')

        for enrollment in enrollments:
            child = enrollment.student
            status = request.POST.get(f'status_{child.id}')
            time_in = request.POST.get(f'time_in_{child.id}') or None
            time_out = request.POST.get(f'time_out_{child.id}') or None
            remarks = request.POST.get(f'remarks_{child.id}', '').strip()

            if status:
                Attendance.objects.update_or_create(
                    child=child,
                    date=form_date,
                    class_obj=form_class,
                    defaults={
                        'status': status,
                        'time_in': time_in,
                        'time_out': time_out,
                        'remarks': remarks,
                        'recorded_by': teacher,
                    }
                )
                saved_count += 1

        messages.success(request, f'✅ Attendance saved for {saved_count} student(s).')
        return redirect(f"{request.path}?date={form_date}&class_id={form_class_id}")

    context = {
        'teacher': teacher,
        'classes': classes,
        'selected_date': selected_date,
        'selected_class': selected_class,
        'children': children,
        'status_choices': Attendance.STATUS_CHOICES,
    }

    return render(request, 'monitoring/record_attendance.html', context)



@login_required
@teacher_required
def student_attendance_detail(request, student_id):
    """View attendance history for a specific student"""
    teacher = request.user.teacher_profile
    student = get_object_or_404(Child, id=student_id)
    
    # Verify access
    if not Enrollment.objects.filter(
        student=student,
        class_obj__teacher=teacher,
        is_active=True
    ).exists():
        messages.error(request, 'You do not have access to this student.')
        return redirect('monitoring:class_list')
    
    # Get date range from query params or default to current month
    end_date = date.today()
    start_date = end_date.replace(day=1)
    
    if request.GET.get('start_date'):
        start_date = datetime.strptime(request.GET.get('start_date'), '%Y-%m-%d').date()
    if request.GET.get('end_date'):
        end_date = datetime.strptime(request.GET.get('end_date'), '%Y-%m-%d').date()
    
    # Get attendance records
    attendance_records = Attendance.objects.filter(
        child=student,
        date__gte=start_date,
        date__lte=end_date
    ).order_by('-date')
    
    # Calculate statistics
    attendance_stats = get_attendance_stats(student, start_date, end_date)
    
    context = {
        'teacher': teacher,
        'student': student,
        'attendance_records': attendance_records,
        'attendance_stats': attendance_stats,
        'start_date': start_date,
        'end_date': end_date,
    }
    
    return render(request, 'monitoring/student_attendance_detail.html', context)


# ========================================
# REPORT GENERATION
# ========================================

@login_required
@parent_required
def download_report_card(request, child_id, quarter):
    """
    Generate and download report card for a child
    This is for PARENTS to download their child's report card
    """
    parent = request.user.parent_profile
    child = get_object_or_404(Child, id=child_id, parents=parent)
    
    # Get class enrollment
    try:
        enrollment = Enrollment.objects.get(student=child, is_active=True)
        class_obj = enrollment.class_obj
    except Enrollment.DoesNotExist:
        messages.error(request, 'Child is not enrolled in any class.')
        return redirect('users:parent_dashboard')
    
    # Get all competency records for this quarter
    domains_data = []
    
    for domain in Domain.objects.prefetch_related('competencies').all():
        competencies_data = []
        
        for competency in domain.competencies.all():
            try:
                record = QuarterlyCompetencyRecord.objects.get(
                    child=child,
                    competency=competency,
                    quarter=quarter
                )
                level = record.level or 'Not Assessed'
            except QuarterlyCompetencyRecord.DoesNotExist:
                level = 'Not Assessed'
            
            competencies_data.append({
                'code': competency.code,
                'description': competency.description,
                'level': level,
            })
        
        if competencies_data:
            domains_data.append({
                'name': domain.name,
                'competencies': competencies_data,
            })
    
    # Get quarterly summary if exists
    try:
        summary = QuarterlySummary.objects.get(
            child=child,
            quarter=quarter,
            class_obj=class_obj
        )
        teacher_remarks = summary.teacher_remarks
    except QuarterlySummary.DoesNotExist:
        teacher_remarks = ''
    
    # Get attendance for this quarter
    # Define quarter date ranges (adjust as needed)
    quarter_dates = {
        1: (date(date.today().year, 6, 1), date(date.today().year, 8, 31)),
        2: (date(date.today().year, 9, 1), date(date.today().year, 11, 30)),
        3: (date(date.today().year, 12, 1), date(date.today().year + 1, 2, 28)),
        4: (date(date.today().year, 3, 1), date(date.today().year, 5, 31)),
    }
    
    start_date, end_date = quarter_dates.get(quarter, (date.today(), date.today()))
    attendance_stats = get_attendance_stats(child, start_date, end_date, class_obj)
    
    # Create Excel report card
    wb = Workbook()
    ws = wb.active
    ws.title = f"Q{quarter} Report Card"
    
    # Styling
    title_font = Font(bold=True, size=16)
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:D1')
    ws['A1'] = f"KINDERGARTEN REPORT CARD - QUARTER {quarter}"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Student Info
    ws['A3'] = "Student Name:"
    ws['B3'] = child.get_full_name()
    ws['A4'] = "LRN:"
    ws['B4'] = child.lrn
    ws['A5'] = "Class:"
    ws['B5'] = class_obj.class_name
    ws['A6'] = "School Year:"
    ws['B6'] = class_obj.school_year
    
    # Make labels bold
    for row in [3, 4, 5, 6]:
        ws[f'A{row}'].font = Font(bold=True)
    
    # Competency Records
    current_row = 8
    ws[f'A{current_row}'] = "COMPETENCY ASSESSMENT"
    ws[f'A{current_row}'].font = header_font
    current_row += 1
    
    # Legend
    ws[f'A{current_row}'] = "Legend: B = Beginning | D = Developing | C = Consistent"
    ws[f'A{current_row}'].font = Font(italic=True, size=10)
    current_row += 2
    
    # Headers
    ws[f'A{current_row}'] = "Competency Code"
    ws[f'B{current_row}'] = "Description"
    ws[f'C{current_row}'] = "Level"
    
    for cell in [ws[f'A{current_row}'], ws[f'B{current_row}'], ws[f'C{current_row}']]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    current_row += 1
    
    # Fill competency data
    for domain_data in domains_data:
        # Domain header
        ws.merge_cells(f'A{current_row}:C{current_row}')
        ws[f'A{current_row}'] = domain_data['name'].upper()
        ws[f'A{current_row}'].font = Font(bold=True, size=11)
        ws[f'A{current_row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left')
        current_row += 1
        
        # Competencies
        for comp in domain_data['competencies']:
            ws[f'A{current_row}'] = comp['code']
            ws[f'B{current_row}'] = comp['description']
            ws[f'C{current_row}'] = comp['level']
            
            # Apply borders
            for col in ['A', 'B', 'C']:
                ws[f'{col}{current_row}'].border = border
            
            ws[f'C{current_row}'].alignment = Alignment(horizontal='center')
            current_row += 1
        
        current_row += 1  # Space between domains
    
    # Attendance Summary
    current_row += 1
    ws[f'A{current_row}'] = "ATTENDANCE SUMMARY"
    ws[f'A{current_row}'].font = header_font
    current_row += 2
    
    ws[f'A{current_row}'] = "Total Days:"
    ws[f'B{current_row}'] = attendance_stats['total_days']
    current_row += 1
    
    ws[f'A{current_row}'] = "Present:"
    ws[f'B{current_row}'] = attendance_stats['present_days']
    current_row += 1
    
    ws[f'A{current_row}'] = "Absent:"
    ws[f'B{current_row}'] = attendance_stats['absent_days']
    current_row += 1
    
    ws[f'A{current_row}'] = "Late:"
    ws[f'B{current_row}'] = attendance_stats['late_days']
    current_row += 1
    
    ws[f'A{current_row}'] = "Attendance Rate:"
    ws[f'B{current_row}'] = f"{attendance_stats['attendance_rate']}%"
    current_row += 2
    
    # Teacher Remarks
    ws[f'A{current_row}'] = "TEACHER'S REMARKS"
    ws[f'A{current_row}'].font = header_font
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:C{current_row + 3}')
    ws[f'A{current_row}'] = teacher_remarks if teacher_remarks else "No remarks provided."
    ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws[f'A{current_row}'].border = border
    
    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 70
    ws.column_dimensions['C'].width = 15
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="Report_Card_{child.get_full_name()}_Q{quarter}.xlsx"'
    )
    wb.save(response)
    return response


@login_required
@teacher_required
def class_report(request, class_id):
    """Generate comprehensive class report"""
    teacher = request.user.teacher_profile
    class_obj = get_object_or_404(Class, id=class_id, teacher=teacher)
    
    quarter = int(request.GET.get('quarter', get_current_quarter()))
    
    # Get all enrolled students
    enrollments = Enrollment.objects.filter(
        class_obj=class_obj,
        is_active=True
    ).select_related('student').order_by('student__last_name', 'student__first_name')
    
    # Prepare student data
    students_data = []
    
    for enrollment in enrollments:
        student = enrollment.student
        
        # Get competency records
        records = QuarterlyCompetencyRecord.objects.filter(
            child=student,
            quarter=quarter
        )
        
        total_competencies = Competency.objects.count()
        assessed = records.exclude(level__isnull=True).count()
        beginning = records.filter(level='B').count()
        developing = records.filter(level='D').count()
        consistent = records.filter(level='C').count()
        
        # Get attendance
        attendance_stats = get_attendance_stats(student, class_obj=class_obj)
        
        students_data.append({
            'student': student,
            'total_competencies': total_competencies,
            'assessed': assessed,
            'beginning': beginning,
            'developing': developing,
            'consistent': consistent,
            'attendance_rate': attendance_stats['attendance_rate'],
        })
    
    context = {
        'teacher': teacher,
        'class_obj': class_obj,
        'quarter': quarter,
        'students_data': students_data,
        'quarters': [1, 2, 3, 4],
    }
    
    return render(request, 'monitoring/class_report.html', context)


@login_required
@teacher_required
def export_class_report(request, class_id):
    """Export class report to Excel"""
    teacher = request.user.teacher_profile
    class_obj = get_object_or_404(Class, id=class_id, teacher=teacher)
    
    quarter = int(request.GET.get('quarter', get_current_quarter()))
    
    # Get enrolled students
    enrollments = Enrollment.objects.filter(
        class_obj=class_obj,
        is_active=True
    ).select_related('student').order_by('student__last_name', 'student__first_name')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Class Report"
    
    # Styling
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = f"CLASS REPORT - {class_obj.class_name} - Quarter {quarter}"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = [
        'LRN', 'Student Name', 'Total Competencies', 'Assessed',
        'Beginning', 'Developing', 'Consistent', 'Attendance Rate'
    ]
    
    ws.append([])  # Empty row
    ws.append(headers)
    
    header_row = ws[3]
    for cell in header_row:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Data rows
    for enrollment in enrollments:
        child = enrollment.student
        
        # Get competency stats
        records = QuarterlyCompetencyRecord.objects.filter(
            child=child,
            quarter=quarter
        )
        
        total_competencies = Competency.objects.count()
        assessed = records.exclude(level__isnull=True).count()
        beginning = records.filter(level='B').count()
        developing = records.filter(level='D').count()
        consistent = records.filter(level='C').count()
        
        # Get attendance
        attendance_stats = get_attendance_stats(child, class_obj=class_obj)
        
        ws.append([
            child.lrn,
            child.get_full_name(),
            total_competencies,
            assessed,
            beginning,
            developing,
            consistent,
            f"{attendance_stats['attendance_rate']}%"
        ])
    
    # Apply borders to all data cells
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=8):
        for cell in row:
            cell.border = border
            if cell.column in [3, 4, 5, 6, 7]:
                cell.alignment = Alignment(horizontal='center')
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 16
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="Class_Report_{class_obj.class_name}_Q{quarter}.xlsx"'
    )
    wb.save(response)
    return response